import os
import pandas as pd
import asyncio
import re
from playwright.async_api import async_playwright
from openpyxl import load_workbook
from openpyxl.drawing.image import Image

# Load URLs from Excel file
file_path = "USAcomplete.xlsx"  # Update with your file path
df = pd.read_excel(file_path)

# Clean column headers and check required columns
df.columns = df.columns.str.strip()
url_column = "WEBSITE URL"  # Adjust if needed
city_column = "CITY"  # Adjust if needed

# Ensure URLs and City names are strings
df[url_column] = df[url_column].astype(str).str.strip()
df[city_column] = df[city_column].astype(str).str.strip()

# Remove empty URLs
df = df[df[url_column].notna()]
urls = df[url_column].tolist()

# Output directory for screenshots
output_dir = "screenshots"
os.makedirs(output_dir, exist_ok=True)

# Function to extract clean domain name (without www and .com)
def clean_domain(url):
    url = url.replace("http://", "").replace("https://", "").split("/")[0]
    url = re.sub(r"^www\.", "", url)  # Remove 'www.'
    url = re.sub(r"\.\w+$", "", url)  # Remove '.com', '.net', '.org', etc.
    return url

# Function to sanitize filenames
def sanitize_filename(url, city):
    domain = clean_domain(url)
    safe_city = city.replace(" ", "_").replace("/", "_")
    return os.path.join(output_dir, f"{domain}_{safe_city}.png")  # Format: domain_city.png

# Asynchronous function to capture a single screenshot
async def capture_screenshot(url, city, retries=3):
    screenshot_path = sanitize_filename(url, city)

    for attempt in range(retries):
        try:
            async with async_playwright() as p:
                browser = await p.chromium.launch(headless=True)
                page = await browser.new_page()

                # Ensure valid URL
                if not url.startswith(("http://", "https://")):
                    url = "https://" + url

                print(f"🌍 Visiting: {url} (Attempt {attempt + 1})")

                # Set a real browser user-agent
                await page.set_extra_http_headers({"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"})

                # Load page with no timeout (wait as long as needed)
                await page.goto(url, timeout=0, wait_until="domcontentloaded")

                # Wait for main content to load
                await page.wait_for_selector("body")  # Ensures the page is visible
                await page.wait_for_load_state("networkidle")  # Ensures all assets have loaded

                # Capture only the visible viewport (not full page)
                await page.screenshot(path=screenshot_path)

                await browser.close()
                print(f"✅ Screenshot saved: {screenshot_path}")

                return screenshot_path  # Return successful screenshot path

        except Exception as e:
            print(f"⚠️ Error capturing {url} (Attempt {attempt + 1}): {e}")

    print(f"❌ Failed to capture {url} after {retries} attempts.")
    return None  # Return None if all retries fail

# Async function to process screenshots in batches of 10
async def capture_all_screenshots():
    batch_size = 10  # Process 10 at a time
    screenshot_results = []

    for i in range(0, len(df), batch_size):
        print(f"🚀 Processing batch {i//batch_size + 1} of {len(df)//batch_size + 1}...")
        tasks = [capture_screenshot(row[url_column], row[city_column]) for _, row in df.iloc[i:i+batch_size].iterrows()]
        results = await asyncio.gather(*tasks)
        screenshot_results.extend(results)

    return screenshot_results

# Run the async function
screenshot_results = asyncio.run(capture_all_screenshots())

# Store screenshot paths in DataFrame
df["Screenshots"] = screenshot_results

# Save updated Excel file (temporary without images)
df.to_excel(file_path, index=False)

# 🖼️ Insert images into Excel
wb = load_workbook(file_path)
ws = wb.active

# Find column index for "Screenshots"
screenshot_col = ws.max_column + 1
ws.cell(row=1, column=screenshot_col, value="Embedded Screenshot")  # Header

for i, row in enumerate(df.itertuples(), start=2):  # Start from row 2 (skip header)
    screenshot_path = getattr(row, "Screenshots", None)
    
    if screenshot_path and os.path.exists(screenshot_path):
        img = Image(screenshot_path)
        img.width, img.height = 200, 150  # Resize image
        img.anchor = f"{chr(65 + screenshot_col)}{i}"  # Example: 'D2', 'D3', etc.
        ws.add_image(img)

# Save Excel file with embedded images
wb.save(file_path)
print(f"🎉 Screenshots embedded in {file_path}")
